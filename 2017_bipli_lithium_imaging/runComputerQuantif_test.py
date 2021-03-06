# -*- coding: utf-8 -*-
"""
Created on Mon Jul  2 12:17:52 2018

@author: JS247994
"""

#function run_Compute_Quantif(Rawdatdir,subject,Processeddir,subjectnumber,Codedir,Pythonexe)

import os, warnings #, sys
import openpyxl
import argparse
import numpy as np
import nibabel as nib

from launch_reconstruct_TPI import launch_reconstruct
from reconstructionTPI import ComputeQuantif_test,nifty_funclib

parser = argparse.ArgumentParser(epilog="RawDaRec (RawDataReconstruction) version 1.0")
parser.add_argument("--rdir", type=str,help="Raw data folder")
parser.add_argument("--pdir", type=str,help="Result (processed) data folder")
parser.add_argument("--subj", type=str,help="Subject identifier")
parser.add_argument("--subjnum", type=str,help="Subject number, assigned to result files")
parser.add_argument("--cdir", type=str, help="The directory used to do the reconstruction (should usually be near working directory")
parser.add_argument("--B0map", type=str, help="Input B0 map path (if absent, set to 1 everywhere)")
parser.add_argument("--v","--verbose", help="output verbosity", action="store_true")
args = parser.parse_args()


cwd = os.getcwd() 
#dir_path = os.path.dirname(os.path.realpath(__file__))

if not args.rdir:
    raise
else:
    Rawdatdir=args.rdir
    
if not args.pdir:
    raise
else:
    Processeddir=args.pdir

if args.v: verbose=True
else: verbose = False

if not args.subj:
    subject="2018_06_29"
else:
    subject=args.subj

if not args.subjnum :
    subjlist = os.listdir(Rawdatdir);
    subjlist.sort();
    i=0;
    for subj in subjlist:
        i=i+1;
        if subject==subj:
            if ( i > 9 ): 
                subjectnumber =  int(i) ;
            else:
                subjectnumber =  [ '0', int(i) ] ;
else:
    subjectnumber=args.subjnum

if not args.cdir:
    Codedir=os.path.join(cwd,'reconstructionTPI');
else:
    Codedir=args.cdir

excelT1s=os.path.join(cwd,'info_pipeline','T1vals.xlsx')
try:
    wb = openpyxl.load_workbook(excelT1s)
    sheet = wb.get_sheet_by_name('Feuil1')
    for l in range(sheet.max_row):
        for m in range(sheet.max_column):
            if ('subj'==sheet[chr(l+64)+str(m+1)].value):
                subjcol=m+1;
                subjstartl=l+2;
            elif ('T1 vals'==sheet[chr(l+64)+str(m+1)].value):
                T1col=m+1;
    for l in range(subjstartl,(sheet.max_row)):
        if (sheet[chr(subjcol+64)+str(l)].value==subject):
            T1valtest=sheet[chr(l+64),str(T1col)].value;
except:
    warnings.simplefilter("default")
        
if not os.path.isdir(Processeddir):
    #os.mkdir(Processeddir)
    os.makedirs(os.path.join(Processeddir,subject))
else:
    if not os.path.isdir(os.path.join(Processeddir,subject)):
        os.mkdir(os.path.join(Processeddir,subject));

Subjectdirr=os.path.join(Rawdatdir,subject);
Subjectdirtwix=os.path.join(Subjectdirr,'twix7T');
Subjectdirp=os.path.join(Processeddir,subject);
Subjectdirresult=os.path.join(Subjectdirp,'TPI','Reconstruct_gridding','01-Raw');

#Create the processeddir subject folders if they are not already created
if not os.path.isdir(os.path.join(Subjectdirp,'Anatomy3T')):
    os.mkdir(os.path.join(Subjectdirp,'Anatomy3T'));
if not os.path.isdir(os.path.join(Subjectdirp,'Anatomy7T')):
    os.mkdir(os.path.join(Subjectdirp,'Anatomy7T'));
if not os.path.isdir(os.path.join(Subjectdirp,'TPI')):
    os.makedirs(os.path.join(Subjectdirp,'TPI','Reconstruct_gridding','01-Raw'));
    os.mkdir(os.path.join(Subjectdirp,'TPI','Reconstruct_gridding','02-PostQuantif'));
    os.mkdir(os.path.join(Subjectdirp,'TPI','Reconstruct_gridding','03-Filtered'));
    os.mkdir(os.path.join(Subjectdirp,'TPI','Reconstruct_gridding','04-7Tanatspace'));
    os.mkdir(os.path.join(Subjectdirp,'TPI','Reconstruct_gridding','05-3Tanatspace'));
    os.mkdir(os.path.join(Subjectdirp,'TPI','Reconstruct_gridding','06-MNIspace'));
if not os.path.isdir(os.path.join(Subjectdirp,'Trufi')):
    os.mkdir(os.path.join(Subjectdirp,'Trufi'));

#Launch the reconstruction of the dat files found in raw to the processed
#images and placed them in right location (should perhaps eventually be
#updated to also include FISTA?
reconstructfile=os.path.join(Codedir,'ProcessData.py');
launch_reconstruct(Subjectdirtwix,Subjectdirresult,subjectnumber,reconstructfile);

filesdirin=os.path.join(Subjectdirp,'TPI','Reconstruct_gridding','01-Raw');
filesdirout=os.path.join(Subjectdirp,'TPI','Reconstruct_gridding','02-PostQuantif');
filesdone=[];
fileslist=os.listdir(filesdirin);
filesnii=[]
for file in fileslist:
    if file.endswith(".nii"):
        filesnii.append(file)

ComputeVFAfile=os.path.join(Codedir,'ComputeDensity3D_clinic.py');
ComputeQuantifile=os.path.join(Codedir,'ComputeQuantif_clinic.py');

T1val=3.947000;

B0Mapfile=os.path.join(Subjectdirp,"BOMap.nii")
maskfile=os.path.join(Subjectdirp,"Mask.nii")

for file1 in filesnii:
    degloc=file1.find('deg');
    over=0;
    for filesdon in filesdone:
        if (file1==filesdon):
            over=1;

    if (degloc)>-1:
        file1path=os.path.join(filesdirin,file1);
        deg1=file1[degloc-2:degloc];   
        if not over:
            for file2 in filesnii:
                VFA= ((file1!=file2) and (file1[degloc:]==file2[degloc:])); #If there is another file with the same name but a different degree value, it is treated as the second VFA file
                if VFA:
                    deg2=file2[degloc-2:degloc-1];
                    Computedniipath=os.path.join(filesdirout,(file1[1:degloc-3]+file1[degloc+4:]));
                    file2path=os.path.join(filesdirin,file2);
                    #codelaunchVFA=({'"'},Pythonexe+{'" '}+ComputeVFAfile+{' --i1 '}+file1path+{' --i2 '}+file2path+{' --deg1 '}+deg1+{' --deg2 '}+deg2+{' --t1 '}+ num2str(T1val)+ {' --v --o '}+Computedniipath);
                    #status = system(codelaunchVFA);
                    
                    filesdone=filesdone+file1+file2;

        Computedniipath=os.path.join(filesdirout,file1);
        ComputeQuantif_test.raw_to_concentration_simp(file1path,deg1,T1val,B0Mapfile,maskfile,Computedniipath,verbose)        
        #codelaunchQuant=strcat({'"'},Pythonexe,{'" '},ComputeQuantifile,{' --i '},file1path,{' --deg '},deg1,{' --t1 '}, num2str(T1val), {' --v --o '},Computedniipath);
        #status = system(codelaunchQuant);

  